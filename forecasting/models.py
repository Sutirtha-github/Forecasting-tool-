from typing import List
from prophet import Prophet
from prophet.serialize import model_to_json
from pmdarima.arima import auto_arima
import time 
from sklearn.metrics import mean_absolute_percentage_error, mean_squared_error, r2_score
import pandas as pd
import matplotlib.pyplot as plt
import random
from itertools import combinations
import numpy as np
from sklearn.feature_selection import f_regression
from sklearn.feature_selection import SelectKBest, SelectPercentile
from sklearn.feature_selection import mutual_info_regression
from statsmodels.stats.outliers_influence import variance_inflation_factor
from scipy.stats import pearsonr
import statsmodels.api as sm
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.model_selection import ParameterGrid
from .utilities import *

def flatten(list_of_lists):
    flattened = []
    for item in list_of_lists:
        flattened.extend(item)
    return flattened

def calculate_mape(actual, predicted):
    merged = actual.rename(columns={'y' : 'actual'}).merge(predicted.rename(columns={'y':'predicted'}), on='ds')
    merged['mape'] = 100*abs(merged['actual'] - merged['predicted']) / merged['actual'].clip(lower=1)
    return merged['mape'].mean(), merged

class ProphetForecaster:
    
    DEFAULT_HYP_PARAMETERS = {'yearly_order' : 25, 'monthly_order' : 3, 'weekly_order' : 2}
    HYP_PARAMETER_SPACE = {'yearly_order' : [0,10,15,20,25], 'monthly_order' : [3],'weekly_order':[2]}
    
    @classmethod
    def disable_all_seasonalities(cls, model):
        """
            Disable all seasonalities in the Prophet model

            Parameters
            ----------
            model
                Prophet model

            Returns
            -------
                Prophet model with seasonalities disabled
        """
        model.weekly_seasonality = False
        model.monthly_seasonality = False
        model.yearly_seasonality = False
        return model
    
    @classmethod
    def create_model(cls,flat_trend : bool, regressors : List[str], param_grid : dict) -> Prophet:
        """
            Creates and returns a prophet model with given parameters

            Parameters
            ----------
            flat_trend
                Boolean to specify if the model should have a flat trend
            regressors
                List of variable names that are to be considered by the prophet
            param_grid
                dictionary with fourier orders for different seasonalities
        """
        if flat_trend:
            model = Prophet(growth = 'flat')
        else:
            model = Prophet()
            
        #remove all default seasonalities
        model = ProphetForecaster.disable_all_seasonalities(model)

        for regressor in regressors:
            model.add_regressor(regressor)
        
        yearly_order = param_grid.get('yearly_order', 0)
        monthly_order = param_grid.get('monthly_order', 0)
        weekly_order = param_grid.get('weekly_order', 0)

        if yearly_order > 0:
            model.add_seasonality('yearly', period = 365.25, fourier_order = yearly_order)
        if monthly_order > 0:
            model.add_seasonality('monthly', period = 30.5, fourier_order = monthly_order)
        if weekly_order > 0:
            model.add_seasonality('weekly', period = 7, fourier_order = weekly_order)

        return model

    @classmethod
    def _fit_model(cls, flat_trend, features, parameters, train, test):
        model = ProphetForecaster.create_model(flat_trend, features, parameters)
        model.fit(train)
        train_pred, test_pred = ProphetForecaster.make_predictions(model, train, test)
        train_mape, train_details = calculate_mape(train, train_pred)
        test_mape, test_details = calculate_mape(test, test_pred)
        train_r2 = r2_score(train['y'].values, train_pred['y'].values)
        test_r2 = r2_score(test['y'].values, test_pred['y'].values)
        train_rmse = mean_squared_error(train['y'].values, train_pred['y'].values)**0.5
        test_rmse = mean_squared_error(test['y'].values, test_pred['y'].values)**0.5
        return {
            'Model' : model,
            'Train MAPE' : np.round(train_mape,2),
            'Test MAPE' : np.round(test_mape,2),
            'Train R2' : np.round(train_r2,2),
            'Test R2' : np.round(test_r2,2),
            'Train RMSE' : np.round(train_rmse,2),
            'Test RMSE' : np.round(test_rmse,2),
            'Parameters' : parameters,
            'Features' : features,
            'Flat Trend' : flat_trend
        }
    
    @classmethod
    def make_predictions(cls,model, *tests : pd.DataFrame) -> List[pd.DataFrame]:
        """
            Generate predictions using the model for all dates in the provided tests

            Parameters
            ---------
            model
                Prophet model to be used for making predictions
            tests
                DataFrame to be used for maing predictions

            Returns
            -------
                List of prediction dataframes
        """
        if len(tests) == 0:
            raise AttributeError("provide atleast one test dataframe")
        predictions = []
        for test in tests:
            pred = model.predict(test).rename(columns={'yhat':'y'})
            pred['y']  = pred.y.clip(0)
            predictions.append(pred.copy())
        return predictions

    @classmethod
    def predict(cls,model, prediction_df):
        return ProphetForecaster.make_predictions(model, prediction_df)[0]
    
    @classmethod
    def fit(cls,train,test,features,hyper_parameters, **kwargs):
        flat_trend = kwargs.get('Flat Trend')
        use_dummies = kwargs.get('use_dummies',[])
        if flat_trend is None:
            results = []
            for flat_trend in [True, False]:
                res = ProphetForecaster._fit_model(flat_trend, features, hyper_parameters, train, test)
                results.append(res)
        else:
            results = ProphetForecaster._fit_model(flat_trend, features, hyper_parameters, train, test)    
        return results
        
        
class ARIMAForecaster:
     
    DEFAULT_HYP_PARAMETERS = {'start_p':0, 'd':1, 'start_q':0,
                                'max_p':6, 'max_d':3, 'max_q':6,
                                'start_P':0, 'D':1, 'start_Q':0,
                                'max_P':6, 'max_D':3, 'max_Q':6}
    
    HYP_PARAMETER_SPACE = {}
    def __init__(self, date_freq):
        self.date_freq = date_freq
    
    def predict(self,model, data):
        data_endo = data['y']
        data_exo = data[self.features] if len(self.features) > 0 else None
        print(self.features)
        print(data_exo)
        predictions = model.predict(n_periods = data_endo.shape[0],X=data_exo)
        return predictions.to_frame(name='y')
        
    def create_model(self, train_endo, train_exo,hyper_params):
    
        # 'm' parameter relates to the number of observations per seasonal cycle, must be known apriori

        if self.date_freq=='M':
            m=12 # monthly
        elif self.date_freq=='Q':
            m=4 #quarterly
        elif self.date_freq=='A':
            m=1 # annualy
        elif self.date_freq=='W':
            m=52 # weekly
        else:
            m=7 # daily


        """
        p = Number of significant terms in PACF for seasonality

        d = Order of differencing for seasonality

        q = Number of significant terms in ACF for seasonality

        P = Number of significant terms in PACF for trend

        D = Order of differencing for trend

        Q= Number of significant terms in ACF for trend

        S= Number of seasonal periods

        """
        arima_model = auto_arima(y = train_endo, X = train_exo,
                                **hyper_params,
                                m=m,
                                seasonal=True, suppress_warnings=True,
                                error_action='warn', trace=False, random_state=42, n_fits=20)

        return arima_model
        
    def fit(self,train,test,feature_space, hyper_parameters, **kwargs):
        self.features = feature_space
        train_endo = train['y']
        train_exo = train[feature_space] if len(feature_space) >0 else None
        model = self.create_model(train_endo, train_exo,hyper_parameters)
        
        predictions_train = self.predict(model,train)
        predictions_test = self.predict(model,test)
        train_mape = 100*mean_absolute_percentage_error(train_endo.to_numpy(), predictions_train)
        test_mape = 100*mean_absolute_percentage_error(test.y.to_numpy(), predictions_test)
        train_r2 = r2_score(train_endo.to_numpy(), predictions_train)
        test_r2 = r2_score(test.y.to_numpy(), predictions_test)
        train_rmse = mean_squared_error(train_endo.to_numpy(), predictions_train)**0.5
        test_rmse = mean_squared_error(test.y.to_numpy(), predictions_test)**0.5
        return {
            'Model' : model,
            'Train MAPE' : np.round(train_mape,2),
            'Test MAPE' : np.round(test_mape,2),
            'Train R2' : np.round(train_r2,2),
            'Test R2' : np.round(test_r2,2),
            'Train RMSE' : np.round(train_rmse,2),
            'Test RMSE' : np.round(test_rmse,2),
            'Parameters' : hyper_parameters,
            'Features' : feature_space
        }
        
    def model_performance(self,train, test, predictions):
        prediction_list = predictions['y'].tolist()
        test_list = test['y'].tolist()

        # Calculate different metrics like MAPE, RMSE, R2
        fig,ax = plt.subplots(figsize=(15,6))
        mape = mean_absolute_percentage_error(test_list, prediction_list)
        rmse = mean_squared_error(test_list, prediction_list)**0.5
        r2 = r2_score(test_list, prediction_list)

        # Plot the time series data and the model predictions
        fig,ax = plt.subplots(figsize=(15,6))
        plt.plot(train.index, train.tolist(),label='Train')
        plt.plot(test.index, test.tolist(),label='Test')
        plt.plot(test.index, predictions.tolist(),label='Predictions')
        plt.legend()
        plt.show()

        print('MAPE:', mape)
        print('RMSE:', rmse)
        print('R2 score:', r2)
        
        
class Forecaster:
    def __init__(self,date_column,output_column, feature_transform_dict,algorithm,date_freq,dummy_columns = []):
        """
            date_column : name of the date column in the data [str]
            output_column : name of the output column in the data [str]
            feature_transform_dict : Dictionary with features as keys and tranforamtion dict as values [dict]
            algorithm : name of the algorithm to be used [str] [prophet|ARIMA|LR]
            data_freq : Date frequency at which model has to be built [str] [required for sarima]
            use_dummies : list of the dummies necessary [list]
        """
        self.date_column = date_column
        self.output_column = output_column
        self.feature_transform_dict = feature_transform_dict
        self.feature_space = Forecaster.create_feature_space_dict(self.feature_transform_dict)
        self.algorithm = algorithm
        self.date_freq = date_freq
        self.dummy_columns = dummy_columns
        if self.algorithm == 'prophet':
            self.model_class = ProphetForecaster() 
        elif self.algorithm == 'arima':
            self.model_class = ARIMAForecaster(self.date_freq)
        elif self.algorithm == 'lr':
            self.model_class = LRForecaster()
        else:
            raise ValueError
        
    
    def transform_data(self,data):
            return data.rename(columns = {self.date_column : 'ds', self.output_column : 'y'})
        
        
    def fit(self, train, test, num_iterations, train_contrib=0.4, test_contrib=0.6):
        
        feature_space = Forecaster.get_random_samples(self.feature_space)
    
        if len(feature_space) == 0:
            feature_space = [[]]
            
        if num_iterations < len(feature_space):
            feature_space = feature_space[:num_iterations]
            
        train = self.transform_data(train)
        test = self.transform_data(test)
        st = time.time()
        
        ## feature selection
        feature_selection_results = []
        
        for feature_combination in feature_space:
            if len(self.dummy_columns) > 0 :
                feature_combination.extend(self.dummy_columns)
            res = self.model_class.fit(train, test,feature_combination, self.model_class.DEFAULT_HYP_PARAMETERS)
            if isinstance(res, list):
                feature_selection_results.extend(res)
            else:
                feature_selection_results.append(res)
        feature_results_df = pd.DataFrame(feature_selection_results)
        
        feature_results_df['Combined MAPE'] = np.round(train_contrib * feature_results_df['Train MAPE'] +test_contrib * feature_results_df['Test MAPE'],2)
        feature_results_df['MAPE Difference'] = np.round(abs(feature_results_df['Test MAPE'] - feature_results_df['Train MAPE']),2) 
        feature_results_df = feature_results_df.sort_values(by='Combined MAPE')
        
        best_row = feature_results_df.iloc[0]
        
        best_features = best_row['Features']
        best_combined_mape = best_row['Combined MAPE']
        self.best_model = best_row['Model']
        best_train_mape, best_test_mape = best_row['Train MAPE'], best_row['Test MAPE']
        self.best_features = best_row['Features']
        best_params = best_row['Parameters']
        
        #if self.algorithm!='lr':
            #name = self.algorithm.capitalize()
       # else:
            #name = 'Linear Regression'
        #save feature results in excel
        wb = load_workbook('summary.xlsx')
        ws= wb.create_sheet(title=self.algorithm.capitalize()+' Feature Results')
        rows = dataframe_to_rows(feature_results_df.drop('Model',axis=1), index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, dict) or isinstance(value, list) or isinstance(value, object):
                    value  = str(value)
                ws.cell(row=r_idx+2, column=c_idx+1, value=value)
        wb.save('summary.xlsx')
        format_sheet2(feature_results_df.drop('Model',axis=1),'summary.xlsx',self.algorithm.capitalize()+' Feature Results' )
        
        kwargs = {}
        if self.algorithm == 'prophet':
            kwargs = {'Flat Trend' : best_row['Flat Trend']}
        ## hyper parameter selection
        hyp_results = []
        hyp_space = list(ParameterGrid(self.model_class.HYP_PARAMETER_SPACE))
        if len(hyp_space) > 1 or (len(hyp_space) == 1 and len(hyp_space[0]) > 0):
            for hyp in hyp_space:
                res = self.model_class.fit(train, test,self.best_features,hyp,**kwargs)
                hyp_results.append(res)
            
            
            hyp_results_df = pd.DataFrame(hyp_results)
            
            hyp_results_df['Combined MAPE'] = np.round(train_contrib * hyp_results_df['Train MAPE'] + \
                                                test_contrib * hyp_results_df['Test MAPE'] ,2)
            hyp_results_df['MAPE Difference'] = np.round(abs(hyp_results_df['Test MAPE'] - hyp_results_df['Train MAPE']),2)
            hyp_results_df = hyp_results_df.sort_values(by='Combined MAPE')
            
            if feature_results_df.iloc[0]['Combined MAPE'] >= hyp_results_df.iloc[0]['Combined MAPE']:
                best_row = hyp_results_df.iloc[0]
            
            best_combined_mape = best_row['Combined MAPE']
            self.best_model = best_row['Model']
            best_train_mape, best_test_mape = best_row['Train MAPE'], best_row['Test MAPE']
            self.best_features = best_row['Features']
            best_params = best_row['Parameters']
            
            #save hyperparameter results in excel
            rows = dataframe_to_rows(hyp_results_df.drop('Model',axis=1), index=False)
            wb = load_workbook('summary.xlsx')
            ws= wb.create_sheet(title=self.algorithm.capitalize()+' Hyperparameter Results')
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if isinstance(value, dict) or isinstance(value, list) or isinstance(value, object):
                        value  = str(value)
                    ws.cell(row=r_idx+2, column=c_idx+1, value=value)
            wb.save('summary.xlsx')
            format_sheet2(hyp_results_df.drop('Model',axis=1),'summary.xlsx',self.algorithm.capitalize()+' Hyperparameter Results' )
            
        et = time.time()
        
        print('time taken:', et-st)
        
        if self.algorithm != 'prophet':
            self.model_class.features = self.best_features
        
        self.model_performance(train, test)
            
    def predict(self, data):
        data = self.transform_data(data)
        return self.model_class.predict(self.best_model, data)
    
    def model_performance(self,train, test):
        predictions = self.predict(test)
        # Calculate different metrics like MAPE, RMSE, R2
        mape = 100*mean_absolute_percentage_error(test['y'].to_numpy(), predictions['y'].to_numpy())
        rmse = mean_squared_error(test['y'].to_numpy(), predictions['y'].to_numpy())**0.5
        r2 = r2_score(test['y'].to_numpy(), predictions['y'].to_numpy())

        # Plot the time series data and the model predictions
        

        print('MAPE:', mape)
        print('RMSE:', rmse)
        print('R2 score:', r2)
    
        predictions_plot(train, test, predictions)
        test_array = np.array(test.y)
        pred_array = np.array(predictions.y)
        test['residuals'] = test_array - pred_array
        QQ_Plot(test_array - pred_array)
        RvF_plot(test_array, pred_array)
        IQR_resid_outlier_plot(test['residuals'], test.y, threshold=1.5)
        #if self.algorithm!='lr':
            #name = self.algorithm.capitalize()
        #else:
            #name = 'Linear Regression'
        excel_output(['Actual vs pred.png', 'Normal QQ plot.png', 'Residual vs fitted.png', 'Residual Outliers.png'], 'summary.xlsx', self.algorithm.capitalize()+' Diagnostic Plots')
        for file in ['Actual vs pred.png', 'Normal QQ plot.png', 'Residual vs fitted.png', 'Residual Outliers.png']:
            if os.path.exists(file):
                os.remove(file)
        print(f'Diagnostic Plots.xlsx file has been generated and saved to {getcwd()}')
    
    
    
    @classmethod
    def flatten(cls, list_of_lists):
        flattened = []
        for item in list_of_lists:
            flattened.extend(item)
        return flattened
    
    @classmethod
    def create_feature_space_dict(cls, input_feature_dict):
        TRANSFROM_LIST = ['Lags','Moving Averages']
        feature_space = {}
        for feature, transform_dict in input_feature_dict.items():
            feature_combinations = []
            total_feature_transformations = []
            for transform_type in TRANSFROM_LIST:
                if transform_type=='Lags':
                    if transform_type in transform_dict and isinstance(transform_dict[transform_type],list):
                        total_feature_transformations.extend([f'{feature}_lags_{transform}' for transform in transform_dict[transform_type]])
                else:
                    if transform_type in transform_dict and isinstance(transform_dict[transform_type],list):
                        total_feature_transformations.extend([f'{feature}_MA_{transform}' for transform in transform_dict[transform_type]])
            if transform_dict['Use actual?'] == 1:
                total_feature_transformations.append(feature)
            for i in range(transform_dict['# Min transforms'], transform_dict['# Max transforms']+1):
                    feature_combinations.extend(list(combinations(total_feature_transformations,i)))
            feature_space[feature] = feature_combinations
        return feature_space
    
    @classmethod
    def get_random_samples(cls, col_dict, num_samples=int(1e3)) :
        samples = []
        choices = set()
        n = len(col_dict)
        max_iterations = 1.
        if len(col_dict) == 0:
            return samples
        for value in  col_dict.values():
            max_iterations *= len(value)
        num_samples = min(num_samples, max_iterations)
        
        while len(samples) < num_samples:
            indices = [random.randint(0,len(l)-1) for l in col_dict.values()]
            hash_ = int(''.join([str(i) for i in indices]))
            if hash_ not in choices:
                choices.add(hash_)
                samples.append(flatten([l[indices[idx]] for idx, l in enumerate(col_dict.values())]))

        return samples
    
    
    
    

class LRForecaster:
    DEFAULT_HYP_PARAMETERS = {}
    
    HYP_PARAMETER_SPACE = {}
    def __init__(self, uni_flag=False, VIF=5):
        self.uni_flag = uni_flag
        self.VIF = VIF
    
    
    @classmethod
    def drop_var(cls,df,stats_df,Y):
        """ Method to drop categorical variables """
        cat_cols=list(stats_df['feature'][stats_df['type'] == 'object'])
        #drop_var=cat_cols +[Y.name]
        
        cat_cols=list(stats_df['feature'][stats_df['type'] == 'object']) 
        #drop_var=cat_cols +[Y.name]
        drop_var=cat_cols
        data = df[[x for x in list(df.columns) if x not in list(drop_var)]] 
        return data
    
    @classmethod
    def missing_stat(cls,df):
        # Obtain various statistics for the features
        missing_stats = []
        for col in df.columns:
            missing_stats.append((col, df[col].nunique(), df[col].isnull().sum() * 100 / df.shape[0], df[col].value_counts (normalize=True, dropna=False).values[0] * 100, df[col].dtype)) 
            stats_df = pd.DataFrame(missing_stats, columns=['feature', 'unique_values', 'percent_missing', 'percent_biggest_cat', 'type']) 
            stats_df.sort_values('percent_missing', ascending=False)
        return stats_df

    @classmethod
    def univariate_reg(cls,DATA,y):
        """univariate regression to select variables having a p-value of less than 0.05(for dummy variables)"""
        univariate = f_regression(DATA.fillna (0),y)
        univariate = pd.Series(univariate[1])
        univariate.index = DATA.columns
        df_uni=univariate.sort_values (ascending=False).reset_index()
        df_uni.rename({'index': 'feature',0:'univariate'},axis=1, inplace=True)
        uni_list=df_uni['feature'][df_uni['univariate']<0.05].to_list()
        uni_df=pd.DataFrame(columns=uni_list,data=DATA)
        return uni_df

    @classmethod
    def mutual_info(cls,Data,y):
        """pearson correlation to select variables having a correlation of greater than 0.1"""  
        mi = Data.apply(lambda x:abs(pearsonr(x,y)[0]), axis=0)
        df_mi=mi.sort_values (ascending=False).reset_index()
        df_mi.rename({'index': 'feature',0: 'Mutual information'},axis=1, inplace=True) 
        mi_list=df_mi['feature'][df_mi['Mutual information']>0.1].to_list() 
        mi_df=pd.DataFrame(columns=mi_list,data=Data)
        return mi_df

    @classmethod
    def mutual_info1(cls,Data,y,a):
        """Select best single variable based on pearson correlation"""
        mi = Data.apply(lambda x:abs(pearsonr(x,y)[0]), axis=0)
        df_mi=mi.sort_values (ascending=False).reset_index()
        df_mi.rename({'index': 'feature',0: 'Mutual information'}, axis=1, inplace=True) 
        if df_mi.shape[0]>=1:
            feature_=df_mi['feature'].iloc[0]
            mi_df=pd.DataFrame(Data[feature_])
        else:
            mi_df=pd.DataFrame()
        return mi_df

    @classmethod
    def univariate_reg1(cls, data,y,k):
        """Select best single variable based on p-value"""
        univariate = f_regression(data.fillna (0), y)
        univariate = pd.Series(univariate[1])
        univariate.index = data.columns
        univariate=univariate.sort_values (ascending=False)
        if k<data.shape[1]:
            sel_ = SelectKBest(f_regression, k=k).fit(data.fillna(0), y)
        else:
            sel_ = SelectKBest (f_regression, k=data.shape[1]).fit(data.fillna(0), y) 
        uni_df=pd.DataFrame(columns=data.columns[sel_.get_support()], data=data)
        return uni_df

    @classmethod
    def VIF_check(cls,results1,final,Y):
        """Obtain variables sorted in terms of VIF, mutual information and univariate p-values """
        vif_check=pd.DataFrame(columns=results1,data=final)
        vif_data = pd.DataFrame()
        vif_data["feature"] = vif_check.columns
        vif_data["VIF"] = [variance_inflation_factor (vif_check.values, i) for i in range(len(vif_check.columns))] 
        mi = vif_check.apply(lambda x:abs(pearsonr(x, Y)[0]), axis=0).T
        df_mi=mi.sort_values(by=0,ascending=False).reset_index()
        df_mi.rename({'index': 'feature',0: 'Mutual information'},axis=1,inplace=True)
        univariate = f_regression(vif_check.fillna (0), Y)
        
        df_mi.rename({'index': 'feature',0: 'Mutual information'}, axis=1, inplace=True)
        univariate = f_regression(vif_check.fillna(0), Y)
        univariate = pd.Series(univariate[1])
        univariate.index = vif_check.columns
        df_uni=univariate.sort_values(ascending=False).reset_index()
        df_uni.rename({'index': 'feature',0:'univariate'},axis=1, inplace=True)
        check=pd.merge(vif_data, df_mi, on='feature', how='left')
        check1=pd.merge(check, df_uni,on="feature", how='left')
        check1['Mutual information_rank']=check1 [ 'Mutual information'].rank (ascending=False)
        check1['univariate_rank']=check1['univariate'].rank (ascending=True)
        check1['VIF_rank' ]=check1['VIF'].rank (ascending=True)
        check1['Rank_average']=(check1['Mutual information_rank']+check1['univariate_rank']+check1['VIF_rank'])/3 
        check1.sort_values('Rank_average')
        check1=check1[check1[ 'VIF'].notna()]
        check1=check1.sort_values(by='Rank_average').reset_index().iloc[:-1]
        return check1

    @classmethod
    def corr_inde(cls, select,Y,x=0.95):
        """Method to filter out highly correlated variables"""
        high_cor_vars = list()
        corr_summary = pd.DataFrame(columns=['var_i', 'var_j', 'corr_ij','corr_i_dv', 'corr_j_dv'])
        c=1
        input_df_cols = list(select.columns)
        for i in input_df_cols:
            c=c+1
            if i not in high_cor_vars:
                colnames_ls = input_df_cols
                colnames_ls = list([x for x in list(colnames_ls) if x not in list (high_cor_vars)])
                colnames_ls.remove(i)   
                for j in colnames_ls:
                    cor_val = pearsonr(select[i],select[j])[0]
                    if abs(cor_val)>x: 
                        cor_var_i = pearsonr(Y, select[i])[0]
                        cor_var_j = pearsonr (Y, select[j])[0]
                        if abs(cor_var_i)>=abs (cor_var_j) and j not in high_cor_vars:  
                            high_cor_vars.append(j)
                            df_temp = pd.DataFrame({"var_i":[i],"var_j":[j],"corr_ij":[cor_val],"corr_i_dv": [cor_var_i],"corr_j_dv": [cor_var_j]})
                            corr_summary = corr_summary.append(df_temp, ignore_index=True) 
                        if abs (cor_var_i)<abs (cor_var_j) and i not in high_cor_vars:
                            high_cor_vars.append(i)
                            df_temp = pd.DataFrame({"var_i":[i],"var_j":[j],"corr_ij":[cor_val],"corr_i_dv": [cor_var_i],"corr_j_dv": [cor_var_j]})
                            corr_summary = corr_summary.append(df_temp, ignore_index=True) 
        final_vars_df=select[[col for col in list(select.columns) if col not in high_cor_vars]]
        return final_vars_df

    @classmethod
    def num_feat_sel(cls, final_vars_df,Y):
        final_vars_df = LRForecaster.mutual_info(final_vars_df,Y)
        if final_vars_df.shape[1]>0:
            final_vars_sel = LRForecaster.univariate_reg(final_vars_df,Y)
        else:
            final_vars_sel=final_vars_df.copy()
        return final_vars_sel

    @classmethod
    def dummy_feat_sel (cls, stats_df,df,Y):
        cat_cols=list(stats_df['feature'][stats_df['type'] == 'object'])
        cat_df=pd.DataFrame(columns=cat_cols,data=df)
        cat_df.reset_index(drop=True, inplace=True)
        if cat_df.shape[1]>0:
            dummy_df=pd.get_dummies (cat_df,drop_first=True)
            dummy_final=LRForecaster.mutual_info(dummy_df,Y)
        else:
            dummy_final=pd.DataFrame()
        return dummy_final

    @classmethod
    def dummy_df(cls, stats_df,df,Y):
        cat_cols=list(stats_df['feature'][stats_df['type'] == 'object'])
        cat_df=pd.DataFrame(columns=cat_cols,data=df)
        cat_df.reset_index(drop=True, inplace=True)
        if cat_df.shape[1]>0:
            dummy_df=pd.get_dummies (cat_df,drop_first=True)
        else:
            dummy_df = pd.DataFrame()
        return dummy_df
    
    @classmethod
    def missing_imp(cls,stats_df,df):
        # Ignore features with missing percentage > 35
        df=df[stats_df['feature'][stats_df[ 'percent_missing' ]<35]]
        df=df.apply(lambda x:x.fillna(x.interpolate()) if x.dtype in ['object', 'datetime64'] else x.fillna(x.value_counts().index[0])) 
        # df=df.apply(Lambda x:x.fillna(x.value_counts().index[0]) if x. dtype=='object')
        df=df.apply(lambda x:x.fillna (method='bfill', axis=0))
        return df
    
    def fit(self, train,test,feature_space,hyper_params): 
        train_dates = train['ds'].tolist()
        test_dates = test['ds'].tolist()
        df = pd.concat([train[['ds','y',*feature_space]], test[['ds','y',*feature_space]]],axis=0).sort_values(by='ds')
        Y = df[['y']]
        df = df.drop(['ds','y'],axis=1)
        stats_df=LRForecaster.missing_stat(df)
        if self.uni_flag:
            uni_df=LRForecaster.drop_var(df, stats_df,Y)
            uni_df = LRForecaster.univariate_reg(uni_df,Y)
            stats_df=LRForecaster.missing_stat(uni_df)
            df=LRForecaster.missing_imp(stats_df,uni_df)
        else:
            df=LRForecaster.missing_imp(stats_df,df)
            
        df=LRForecaster.drop_var(df, stats_df,Y)
        

        if len(df.columns)>1:
            vif_check_df = LRForecaster.VIF_check(df.columns, df, Y)
            dropping_cols = list(vif_check_df['feature'][vif_check_df[ 'VIF']>self.VIF])
            dropping_cols = list(vif_check_df['feature'][vif_check_df[ 'VIF']>self.VIF]) 
            df = df.drop(dropping_cols, axis=1)

        df=LRForecaster.corr_inde(df,Y,x=0.95)
        self.features = df.columns
        
        X_train= train[self.features]
        X_test = test[self.features]
        y_train = train['y']
        y_test = test['y']

        X_train = sm.add_constant(X_train)        
        model = sm.OLS(y_train, X_train)    
        model = model.fit()        
        X_test = sm.add_constant(X_test, has_constant="add") 
        y_pred_train = self.predict(model,train)
        y_pred_test = self.predict(model,test)
        train_mape = 100*mean_absolute_percentage_error(y_train.to_numpy(), y_pred_train.y.to_numpy())
        test_mape = 100*mean_absolute_percentage_error(y_test.to_numpy(), y_pred_test.y.to_numpy())
        train_r2 = r2_score(y_train.to_numpy(), y_pred_train.y.to_numpy())
        test_r2 = r2_score(y_test.to_numpy(), y_pred_test.y.to_numpy())
        train_rmse = mean_squared_error(y_train.to_numpy(), y_pred_train.y.to_numpy())**0.5
        test_rmse = mean_squared_error(y_test.to_numpy(), y_pred_test.y.to_numpy())**0.5
        return {
            'Model' : model,
            'Train MAPE' : np.round(train_mape,2),
            'Test MAPE' : np.round(test_mape,2),
            'Train R2' : np.round(train_r2,2),
            'Test R2' : np.round(test_r2,2),
            'Train RMSE' : np.round(train_rmse,2),
            'Test RMSE' : np.round(test_rmse,2),
            'Parameters' : hyper_params,
            'Features' : self.features
        }
        #print(self.model_performance(train.set_index('ds')['y'],test.set_index('ds')['y'], predictions.set_index('ds')['y']))

    def predict(self, model, data):
        return model.predict(sm.add_constant(data[self.features])).to_frame(name='y')
    
    def model_performance(self,train, test, predictions):
        prediction_list = predictions.tolist()
        test_list = test.tolist()

        # Calculate different metrics like MAPE, RMSE, R2
        mape = mean_absolute_percentage_error(test_list, prediction_list)
        rmse = mean_squared_error(test_list, prediction_list)**0.5
        r2 = r2_score(test_list, prediction_list)

        # Plot the time series data and the model predictions
        fig,ax = plt.subplots(figsize=(15,6))
        plt.plot(train.index, train.tolist(),label='Train')
        plt.plot(test.index, test.tolist(),label='Test')
        plt.plot(test.index, predictions.tolist(),label='Predictions')
        plt.legend()
        plt.show()

        print('MAPE:', mape)
        print('RMSE:', rmse)
        print('R2 score:', r2)