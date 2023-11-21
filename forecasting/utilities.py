import seaborn as sns
import matplotlib.pyplot as plt
import statsmodels.api as sm
import numpy as np
import matplotlib.dates as mdates
import shap
import os
from os import getcwd
from sklearn.linear_model import LinearRegression
from statsmodels.tsa.seasonal import seasonal_decompose
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from mdutils.mdutils import MdUtils
from pylab import rcParams

def tsa_decomposition(df, time_column, output_column, fill_na, default_date_freq):
    rcParams['figure.figsize'] = 12, 8
    plt.rcParams.update({'font.size': 20})
    if fill_na not in ['ffill','bfill','mean','median']:
        raise ValueError('')
    if fill_na in ['ffill','bfill']:
        res = seasonal_decompose(df[[time_column,output_column]].set_index(time_column).asfreq(default_date_freq).fillna(method=fill_na))

    elif fill_na == 'mean':
        fill_value = df[output_column].mean()
        res = seasonal_decompose(df[[time_column,output_column]].set_index(time_column).asfreq(default_date_freq).fillna(fill_value))
    elif fill_na == 'median':
        fill_value = df[output_column].median()
        res = seasonal_decompose(df[[time_column,output_column]].set_index(time_column).asfreq(default_date_freq).fillna(fill_value))
    res.plot()
    
    plt.ioff()
    plt.savefig('decomposition.png')
    
def RvF_plot(y, y_hat):
    rcParams['figure.figsize'] = 20, 20
    plot = plt.figure()
    plt.rcParams.update({'font.size': 25})
    plot.axes[0] = sns.residplot(x=y_hat, y=y,
                            lowess=True,
                            line_kws={'color': 'red'})

    plot.axes[0].set_title('Residuals vs Fitted Plot')
    plot.axes[0].set_xlabel('Fitted values')
    plot.axes[0].set_ylabel('Residuals')
    plt.ioff()
    plt.savefig('Residual vs fitted.png')
    
def QQ_Plot(residuals):
    plt.rcParams.update({'font.size': 25})
    rcParams['figure.figsize'] = 20, 20
    fig = sm.qqplot(np.array(residuals), fit=True, line="45")
    plt.title('Normal QQ Plot')
    plt.ioff()
    plt.savefig('Normal QQ plot.png')
    
def predictions_plot(train, test, predictions):
    fig,ax = plt.subplots(figsize=(18,12))
    plt.rcParams.update({'font.size': 15})
    plt.plot(train.index, train['y'].to_numpy(),label='Train')
    plt.plot(test.index, test['y'].to_numpy(),label='Test')
    plt.plot(test.index, predictions['y'].to_numpy(),label='Predictions')
    plt.xlabel('Time')
    plt.ylabel('Target')
    plt.title('Actual vs Predicted Plot')
    #plt.setp(plt.gca().get_xticklabels(), rotation=60, ha="right")
    plt.legend()
    plt.ioff()
    plt.savefig('Actual vs pred.png')            
    

def outlier_plot(y,dates, outlier_index):
    
    """
    Input:
    y: Target column (Series)
    outlier_index: list of outlier indices
    
    """
    plt.ion()
    rcParams['figure.figsize'] = 18, 14
    plt.rcParams.update({'font.size': 25})
    fig, ax = plt.subplots()
   
    ax.plot(y.index, y.values, color='black', label = y.name)
    ax.scatter(outlier_index, y.loc[outlier_index].values, color='red', label = 'Outliers')
    ax.set_xticklabels(dates)
    ax.set_ylabel(y.name.capitalize())
    ax.set_xlabel('Time')
    # plt.gca().xaxis.set_major_locator(mdates.DayLocator((1,15)))
    # plt.gca().xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    plt.setp(plt.gca().get_xticklabels(), rotation=60, ha="right")
    plt.tight_layout()
    plt.legend()
    plt.ioff()
    plt.savefig('Outliers.png')
    
    
def shap_viz(df, target, model = LinearRegression):        
    """
    df: DataFrame
    target: name of target column (string)
    model: model instance 
    
    """
    y = df[target]
    X = df.drop([target], axis=1)
    estimator = model()
    estimator.fit(X, y)
    n=min(X.shape[0],100)
    Xn = shap.utils.sample(X, n)
    explainer = shap.Explainer(estimator.predict, Xn)
    shap_values = explainer(Xn)
    plt.ion()
    #shap.plots.bar(shap_values)
    shap.plots.beeswarm(shap_values, show=False)
    plt.ioff()
    plt.savefig('SHAP.png')
    
    
# plot outliers along with target during residual analysis after training

def IQR_resid_outlier_plot(y, Y, threshold=1.5):
    
    """
    y: residuals (array)
    Y: target column (Series/Dataframe)
    """
    plt.rcParams.update({'font.size': 10})
    Q1 = np.percentile(y, 25,
                       interpolation = 'midpoint')
    Q3 = np.percentile(y, 75,
                       interpolation = 'midpoint')
    IQR = Q3 - Q1
    upper = y.index[np.where(y >= (Q3+threshold*IQR))]
    lower = y.index[np.where(y <= (Q1-threshold*IQR))]
    fig, ax = plt.subplots(figsize=(8,6))
    outlier_index = list(upper)+list(lower)
    ax.plot(list(y.index), y.values, color='black', label = 'Residuals', linestyle='--')
    ax.scatter(outlier_index,y.loc[outlier_index].values, color='red', label = 'Residual Outlier')
    ax.plot(list(Y.index), Y.values, color='black', label = 'Actual')
    ax.scatter(outlier_index,Y.loc[outlier_index].values, color='red', label = 'Corresponding Target Outlier')
    ax.set_ylabel('Target')
    ax.set_xlabel('Time')
    ax.set_title('Outliers in Residuals and corresponding Target outliers')
    plt.tight_layout()
    plt.legend()
    plt.ioff()
    plt.savefig('Residual Outliers.png')
    
def excel_output(images, file_name, sheet_name):
    """
    images: list of images
    filename: name/path of excel file as you want to save (e.g. 'TEST.xlsx') 
    
    """
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]
        ws.sheet_view.showGridLines = False
    else:
        wb=Workbook()
        ws = wb.worksheets[0]
        ws.sheet_view.showGridLines = False
    idx=0
    for row in range(2,len(images)//2+2):
        for col in range(2,4):
            ws.row_dimensions[row].height = 300
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 90
            img = Image(images[idx])
            img.height = 300
            img.width = 600
            img.anchor = ws.cell(row=row,column=col).coordinate
            ws.add_image(img)
            idx+=1
    wb.save(file_name)
    for image in images:
        if os.path.exists(image):
            os.remove(image)

def format_sheet1(df, path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = False
    c1=ws.cell(row=1,column=6)
    c1.value="Feature Summary"
    c1.font = Font(name="Arial", size=20, color="00000000")
    # Add cell color code legend
    c2=ws.cell(row=1,column=2)
    c2.value='User choice'
    c2.fill  = PatternFill(fill_type='solid', start_color='00CC99FF', end_color='00CC99FF')
    c2.alignment = Alignment(horizontal='center', vertical='center')
    c3=ws.cell(row=1,column=3)
    c3.value='No options'
    c3.fill  = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
    c3.alignment = Alignment(horizontal='center', vertical='center')
    # Apply boundary around the table
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    thick_border_left = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    thick_border_right = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    thick_border_top = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    thick_border_bottom = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    top_left_corner = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    top_right_corner = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    bottom_left_corner = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    bottom_right_corner = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    
    n_features = len(df.columns)
    for row in range(3,n_features+4):
        for col in range(2,19):
            if row==3 or col ==2:
                ws.cell(row=row, column=col).fill  = PatternFill(fill_type='solid', start_color='0099CCFF', end_color='0099CCFF')
            elif (row>=5 and (col==7 or col==11)) or (row==5 and col==9) or (row>5 and col>=12):
                ws.cell(row=row, column=col).fill  = PatternFill(fill_type='solid', start_color='00CC99FF', end_color='00CC99FF')
            else:
                ws.cell(row=row, column=col).fill  = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
                
            if row==3 and col==2:
                ws.row_dimensions[row].height = 30
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 10
            else:
                ws.row_dimensions[row].height = 30
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
            if row==3:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
            else:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=col).border = thin_border
            if row==3:
                ws.cell(row=row, column=col).border = thick_border_top
            if col==2:
                ws.cell(row=row, column=col).border = thick_border_left
            if row==n_features+3:
                ws.cell(row=row, column=col).border = thick_border_bottom
            if col==18:
                ws.cell(row=row, column=col).border = thick_border_right
            if row==3 and col==2:
                ws.cell(row=row, column=col).border = top_left_corner
            if row==3 and col==18:
                ws.cell(row=row, column=col).border = top_right_corner
            if row==n_features+3 and col==2:
                ws.cell(row=row, column=col).border = bottom_left_corner
            if row==n_features+3 and col==18:
                ws.cell(row=row, column=col).border = bottom_right_corner
    # Adding drop down list to cells for options
    dv1 = DataValidation(type="list", formula1='"mean, median, mode, ffill, bfill"', allow_blank=False)
    dv1.prompt = 'Please select from the list'
    dv1.promptTitle = 'Imputation methods'
    ws.add_data_validation(dv1)
    dv2 = DataValidation(type="list", formula1='"mean, median, mode, sum, min, max"', allow_blank=False)
    dv2.prompt = 'Please select from the list'
    dv2.promptTitle = 'Aggragation functions'
    ws.add_data_validation(dv2)
    dv3 = DataValidation(type="list", formula1='",cap"', allow_blank=True)
    dv3.prompt = 'Please select from the list'
    ws.add_data_validation(dv3)
    dv3.add(ws["I5"])
    dv4 = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv4.prompt = 'Please select from the list'
    ws.add_data_validation(dv4)
    for i in range(1, n_features):
        dv1.add(ws["G{}".format(i+4)])
        dv2.add(ws["K{}".format(i+4)]) 
    if n_features>2:
        for i in range(2, n_features):
            dv4.add(ws["L{}".format(i+4)])
            dv4.add(ws["M{}".format(i+4)])  
    wb.save(path)      
    
def format_sheet2(df, path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = False
    c=ws.cell(row=1,column=5)
    c.value= sheet_name
    c.font = Font(name="Arial", size=20, color="00000000")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    thick_border_left = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    thick_border_right = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    thick_border_top = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    thick_border_bottom = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    top_left_corner = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    top_right_corner = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thick'), 
                         bottom=Side(style='thin'))
    bottom_left_corner = Border(left=Side(style='thick'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    bottom_right_corner = Border(left=Side(style='thin'), 
                         right=Side(style='thick'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))
    
    n_features = len(df.columns)
    for row in range(3,df.shape[0]+4):
        for col in range(2,n_features+2):
            if row==3 :
                ws.cell(row=row, column=col).fill  = PatternFill(fill_type='solid', start_color='0099CCFF', end_color='0099CCFF')
            else:
                ws.cell(row=row, column=col).fill  = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
            ws.row_dimensions[row].height = 30
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
            if row==3:
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=col).border = thin_border
            if row==3:
                ws.cell(row=row, column=col).border = thick_border_top
            if col==2:
                ws.cell(row=row, column=col).border = thick_border_left
            if row==df.shape[0]+3:
                ws.cell(row=row, column=col).border = thick_border_bottom
            if col==n_features+1:
                ws.cell(row=row, column=col).border = thick_border_right
            if row==3 and col==2:
                ws.cell(row=row, column=col).border = top_left_corner
            if row==3 and col==n_features+1:
                ws.cell(row=row, column=col).border = top_right_corner
            if row==df.shape[0]+3 and col==2:
                ws.cell(row=row, column=col).border = bottom_left_corner
            if row==df.shape[0]+3 and col==n_features+1:
                ws.cell(row=row, column=col).border = bottom_right_corner
    wb.save(path)
    
def summary_dict():
    if os.path.exists('Summary Dictionary.MD'):
        os.remove('Summary Dictionary.MD')
    mdAcme = MdUtils(file_name='Summary Dictionary')
    mdAcme.new_header(level=1, title='Dictionary')
    mdAcme.write("\n1. variable : Feature name \n2. dtype : datatype of feature \n3. % missing : missing values percentage\n4. missing: if missing values are present in the feature [bool(Y/N)]\n5. imputation : type of imputation method applied; options available - [mean, median, mode, ffill, bfill]\n6. % outliers: percentage of outliers present in target column\n7.outlier_treatment : method to deal with the outliers; options - [blank: keep outliers, cap: replace the extreme points with a moderate value]\n8. importance: displays the contribution percentage of predictor variables to the target\n9. aggregation : club the time-series to higher order intervals; options available: [mean, median, mode, sum, min, max]\n10. input_to_model : choice of predictor variables [bool(Y/N)]\n10. use_actual : choice to use the actual/non-transformed feature [bool(Y/N)]\n11. lags : list of lagged intervals [e.g. 1,2 will produce feature_lag_1, feature_lag_2]\n12. MA : moving average transformation [e.g. 1,2 will produce feature_MA_1, feature_MA_2]\n13. min_transforms : minimum count of transformed features to include in the feature set while modeling\n14. max_transforms : maximum count of transformed features to include in the feature set while modeling\n15. static features : list of features that should always be present in all feature combinations [e.g feature_lag_1, feature_MA_1]")
    mdAcme.create_md_file()