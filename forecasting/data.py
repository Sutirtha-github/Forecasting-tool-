import pandas as pd
import numpy as np
import re
import math
import os
from os import getcwd
import shap
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from .utilities import shap_viz, outlier_plot, tsa_decomposition, excel_output, summary_dict, format_sheet1

class DataLoader:
    default_transform_mapping = {'int' : 'mean','float' : 'mean','datetime' : 'last','object' : 'last'}
    def __init__(self, 
                 data : pd.DataFrame, 
                 date_column : str, 
                 output_column : str,
                 default_date_freq : str,
                 agg_date_freq : str = None,
                 ):
        
        self.input_df = data
        self.date_column = date_column
        self.output_column = output_column
        self.default_date_freq = default_date_freq
        self.agg_date_freq = agg_date_freq
    
    @classmethod 
    def impute_data(cls,col, impuation_input_dict):
        imputation = impuation_input_dict.get(col.name)
        if imputation == 'mean':
            return col.fillna(col.mean())
        elif imputation == 'median':
            return col.fillna(col.median())
        elif imputation == 'mode':
            return col.fillna(col.mode()[0])
        elif imputation == 'ffill':
            return col.fillna(method=imputation)
        elif imputation == 'bfill':
            return col.fillna(method=imputation)   
        else:
            return col
        
    def generate_summary(self):
        #check for nan in datecolumn 
        if (self.input_df[self.date_column].isna().sum()) > 0:
            raise ValueError(f'Null value exists in date column {self.date_column}')

        #convert to datetime and sort by datetime
        self.input_df[self.date_column] = pd.to_datetime(self.input_df[self.date_column])
        self.input_df = self.input_df.sort_values(by=self.date_column)

        ## check for duplicates
        if self.input_df[self.date_column].nunique() != len(self.input_df):
            raise ValueError('Duplicate dates exist in dataframe')
        original_shape = len(self.input_df)
        ##set default frequency
        self.input_df = self.input_df.set_index(self.date_column).asfreq(self.default_date_freq).reset_index().copy()
        date_missing = original_shape != len(self.input_df)
        
        ## Get dtypes, missing values and missing flag
        summary_df = pd.concat([self.input_df.dtypes,self.input_df.isna().sum()],axis=1).reset_index().rename(columns={'index' : 'Variable',
                                                                           0 : 'Data type',
                                                                           1 : 'Missing count(%)'})
        summary_df['Missing count(%)'] = (100*summary_df['Missing count(%)'] / len(self.input_df)).round(2)
        summary_df['Is missing?'] = summary_df['Missing count(%)'] > 0
        summary_df['Is missing?'] = summary_df['Is missing?'].apply(lambda x : 'Y' if x else 'N')
        summary_df['Imputation method'] = summary_df.apply(lambda x : DataLoader.default_transform_mapping[re.match('[a-z]*', str(x['Data type'])).group()] if x['Is missing?'] =='Y' else np.nan,axis=1)
        summary_df['Outlier count(%)']= np.nan
        summary_df['Outlier Treatment']= np.nan
        summary_df['Feature importance(%)']= np.nan
        
        if self.agg_date_freq is not None and self.default_date_freq != self.agg_date_freq:
            summary_df['Aggregation method'] = summary_df['Data type'].apply(lambda x : DataLoader.default_transform_mapping[re.match('[a-z]*', str(x)).group()])
        else:
            summary_df['Aggregation method'] = np.nan
        summary_df['Input to model?'] = 'N'
        summary_df['Use actual?'] = 'N'
        summary_df['Lags'] = np.nan
        summary_df['Moving Averages'] = np.nan
        summary_df['# Min transforms'] = 0
        summary_df['# Max transforms'] = 1
        summary_df['Static features'] = np.nan
        
        #outliers
        self.outlier_index = DataLoader.IQR(self.input_df[self.output_column])
        summary_df.loc[summary_df.Variable == self.output_column,'Outlier count(%)'] = np.round(DataLoader.percentage_oulier(self.input_df[self.output_column], self.outlier_index),2)
        summary_df.loc[summary_df.Variable == self.output_column,'Outlier Treatment'] = np.nan if len(list(self.outlier_index)) == 0 else 'cap'
        
        # feature importance
        if len(self.input_df.columns)>2:
            shap_importance  = DataLoader.mean_shap_values(self.input_df.drop(self.date_column,axis=1).dropna(), self.output_column)
            for index, row in shap_importance.iterrows():
                summary_df.loc[summary_df.Variable == row['Feature'],'Feature importance(%)'] = row['feature importance values']


        summary_df.loc[summary_df.Variable == self.date_column,'Is missing?'] = 'Y' if date_missing else 'N'
        summary_df.loc[summary_df.Variable == self.date_column,'Missing count(%)'] = round(100*(len(self.input_df) - original_shape) /original_shape,2)
        summary_df.loc[summary_df.Variable == self.date_column,'Imputation method'] = np.nan
        summary_df.loc[summary_df.Variable == self.date_column,'Aggregation method'] = np.nan
        
        #save summary csv
        
        if os.path.exists('summary.xlsx'):
            os.remove('summary.xlsx')
        with pd.ExcelWriter('summary.xlsx', mode='w',) as writer:
            summary_df.to_excel(writer, sheet_name='Pre-transform Data Summary', startcol=1, startrow=2)
        format_sheet1(self.input_df, 'summary.xlsx', 'Pre-transform Data Summary' )
        print(f'summary.xlsx file has been generated and saved to {getcwd()}')
        
        #print('Time series decomposition:')
        tsa_decomposition(self.input_df, self.date_column, self.output_column, fill_na='ffill', default_date_freq=self.default_date_freq)
        #print('Shap values:')
        #shap_viz(self.input_df.drop(self.date_column,axis=1).dropna(), self.output_column)
        #print('Outlier plot:')
        outlier_plot(self.input_df[self.output_column],self.input_df[self.date_column].tolist(), self.outlier_index)
        
        # save plots csv
        excel_output(['decomposition.png', 'Outliers.png'], 'summary.xlsx','Time-Series Analysis Plots')
        wb = load_workbook('summary.xlsx')
        ws = wb['Time-Series Analysis Plots']
        c1=ws.cell(row=1, column=2)
        c1.value = 'Seasonal-Trend Decomposition'
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c2=ws.cell(row=1, column=3)
        c2.value = 'Outliers in Target data'
        c2.alignment = Alignment(horizontal='center', vertical='center')
        wb.save('summary.xlsx')
        
        # save readme file for summary dictionary 
        summary_dict()
        print(f'Summary Dictionary.md file has been generated and saved to {getcwd()}')
        
    def transform_data(self,summary_df=None, path=None):
        if summary_df is None:
            if path is None:
                summary_input_df = pd.read_excel('summary.xlsx', sheet_name='Pre-transform Data Summary', skiprows=2, usecols = range(2,18))
            else:
                summary_input_df = pd.read_excel(path)
        else:
            summary_input_df = summary_df.copy()
            
        ##preprocessing
        summary_input_df['Lags'] = summary_input_df['Lags'].apply(lambda x : list(map(int, x.split(','))) if isinstance(x,str) else x if math.isnan(x) else [int(x)])
        summary_input_df['Moving Averages'] = summary_input_df['Moving Averages'].apply(lambda x : list(map(int, x.split(','))) if isinstance(x,str) else x if math.isnan(x) else [int(x)])
        summary_input_df['Static features'] = summary_input_df['Static features'].apply(lambda x : x.split(',') if isinstance(x,str) else x )

        ## impute dataframe
        ## date column
        date_imputation = 'NA'
        if summary_input_df[summary_input_df.Variable == self.date_column]['Is missing?'].item() == 'Y':
            date_imputation = summary_input_df[summary_input_df.Variable == self.date_column]['Imputation method'].item()

        ## input variables
        imputation_input_dict = summary_input_df[~summary_input_df.Variable.isin([self.date_column])].set_index('Variable')['Imputation method'].to_dict()
        transformed_data = self.input_df.set_index(self.date_column).asfreq(self.default_date_freq).apply(lambda x : DataLoader.impute_data(x, imputation_input_dict)).reset_index()

        ##outlier treatment
        outlier_treatment = summary_input_df.loc[summary_input_df['Variable']==self.output_column, 'Outlier Treatment'].item()
        if isinstance(outlier_treatment,str):
            transformed_data[self.output_column] = DataLoader.cap_outliers(transformed_data[self.output_column], self.outlier_index)
        
        ## aggregation
        if self.agg_date_freq is not None and self.default_date_freq != self.agg_date_freq:
            agg_dict = summary_input_df.set_index('Variable')['Aggregation method'].to_dict()
            del agg_dict[self.date_column]
            assert len(agg_dict) == len(summary_input_df)-1
            transformed_data = transformed_data.set_index(self.date_column).resample(self.agg_date_freq).agg(agg_dict)
                    
        # get transforms from user
        feature_transform_input = summary_input_df[summary_input_df['Input to model?'] == 'Y'].set_index('Variable')[['Use actual?','Lags','Moving Averages','# Min transforms','# Max transforms','Static features']].T.to_dict()

        for feature, transform_dict in feature_transform_input.items():
            if isinstance(transform_dict['Lags'],list):
                transformed_data = transformed_data.assign(**{f'{feature}_lags_{lag}' : transformed_data[feature].shift(lag) for lag in transform_dict['Lags']}).fillna(method='bfill')
                
            if isinstance(transform_dict['Moving Averages'],list):
                transformed_data = transformed_data.assign(**{f'{feature}_MA_{ma}' : transformed_data[feature].rolling(ma).mean() for ma in transform_dict['Moving Averages']}).fillna(method='bfill')
        
        transformed_data = transformed_data.reset_index()
        self.transformed_data = transformed_data
        self.feature_transform_input = feature_transform_input
        
        rows = dataframe_to_rows(transformed_data, index=False)
        
        #save transformed data 
        wb = load_workbook('summary.xlsx')
        ws= wb.create_sheet(title='Transformed Data')
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, dict) or isinstance(value, list) or isinstance(value, object):
                    value  = str(value)
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save('summary.xlsx')
        
        return {
            'date_column' : self.date_column,
            'output_column' : self.output_column,
            'data' : transformed_data,
            'feature_transform_input' : feature_transform_input
        }
    
    @classmethod
    def create_time_dummies(cls,df,time_column,**kwargs):
        df_ = df.copy()
        
        month = kwargs.get('month',False)
        year = kwargs.get('year',False)
        week = kwargs.get('week',False)
        day_of_week = kwargs.get('day_of_week',False)
        quarter = kwargs.get('quarter',False)
        hour = kwargs.get('hour',False)
        
        if month:
            df_ = pd.concat([df_, pd.get_dummies(df[time_column].dt.month,prefix='month')],axis=1)
        if year:
            df_ = pd.concat([df_, pd.get_dummies(df[time_column].dt.isocalendar().year,prefix='year')],axis=1)
        if week:
            df_ = pd.concat([df_, pd.get_dummies(df[time_column].dt.isocalendar().week,prefix='week')],axis=1)
        if day_of_week:
            df_ = pd.concat([df_, 
                            pd.get_dummies(df[time_column].dt.weekday).rename(columns = {idx:col for idx,col in enumerate(['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
                                                                                                                        )})],axis=1)
        if quarter:
            df_ = pd.concat([df_, pd.get_dummies(np.ceil(df[time_column].dt.month / 3).astype(int),prefix='quarter')],axis=1)
        if hour:
            df_ = df_ = pd.concat([df_, pd.get_dummies(df[time_column].dt.hour, prefix='hour')],axis=1)
        
        return df_,list(set(df_.columns) - set(df.columns))
    
    @classmethod
    def mean_shap_values(cls, df, target, model = LinearRegression):
    
        """
        df: DataFrame
        target: name of target column (string)
        model: model instance 
        
        """
        y = df[target]
        X = df.drop([target], axis=1)
        estimator = model()
        estimator.fit(X, y)
        n = min(df.shape[0], 100)
        Xn = shap.utils.sample(X, n)
        explainer = shap.Explainer(estimator.predict, Xn)
        shap_values = explainer(Xn)
        feature_names = X.columns
        shap_df = pd.DataFrame(data=shap_values.values, columns = feature_names)
        vals = np.abs(shap_df).mean(0)
        perc_vals  = [np.round(val*100/sum(vals),1) for val in vals]
        shap_importance = pd.DataFrame(list(zip(feature_names, perc_vals)),
                                        columns=['Feature','feature importance values'])
        return shap_importance  
        
    @classmethod
    def IQR(cls, y, threshold=1.5):
        """
        Input:
        y: target column (DataFrame/Series)
        threshold: float
        
        Output:
        outlier_index: list of outlier indices
        
        """
        Q1 = y.quantile(q=0.25)
        Q3 = y.quantile(q=0.75)
        # Q1 = np.percentile(y, 25,
        #                 interpolation = 'midpoint')
        # Q3 = np.percentile(y, 75,
        #                 interpolation = 'midpoint')
        IQR = Q3 - Q1
        upper = y.index[np.where(y >= (Q3+threshold*IQR))]
        lower = y.index[np.where(y <= (Q1-threshold*IQR))]
        outlier_index = list(upper)+list(lower)
        return outlier_index
    
    @classmethod
    def percentage_oulier(cls, y, outlier_index):
    
        """
        Input:
        y: Target column (Series)
        outlier_index: list of outlier indices
        
        """
        return len(outlier_index)*100/y.shape[0]
    
    @classmethod
    def cap_outliers(cls, target, outlier_index):
    
        """
        Input:
        y: Target column (Series)
        outlier_index: list of outlier indices
        
        Output: new target column with capped outliers (Series/Dataframe)
        
        """
        drop_outliers_df  = target.drop(outlier_index)
        cap_val = max(drop_outliers_df.values)
        for idx in outlier_index:
            target.at[idx] = cap_val
            
        return target
    
   
        